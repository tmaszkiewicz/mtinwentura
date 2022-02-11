from decimal import Context
from django.http.response import HttpResponseRedirect
from django.shortcuts import redirect, render
from django.http import HttpResponse
from .models import indeks, indeks_bazowy,fileform,PlikDoPobrania,plik
from .forms import indeksForm
from .functions import readxls
from django.views.decorators.csrf import csrf_exempt
# Create your views here.
def mtinw_sc(request, *args, **kwargs):
    barcode=""
    nazwa=""
    context = {}
    if request.method == 'POST':
        form = indeksForm(request.POST)
        if form.is_valid():
            barcode = form.cleaned_data['barcode']
            #barcode = request.POST['barcode']
            indeks.objects.create(barcode=barcode,nazwa=nazwa)
    
        context['form'] = form
    context['barcode']=barcode
    url='mtinwentura/mtinw_sc.html'
    return render(request,url,context)
def mtinw_pc(request, *args, **kwargs):
    context = {}
    url='mtinwentura/mtinw_pc.html'
    indeksy = indeks.objects.all()
    context['indeksy']= indeksy
    
    return render(request,url,context)

def mtinw_pc2(request, *args, **kwargs):
    context = {}
    url='mtinwentura/mtinw_pc2.html'
    indeksy = indeks_bazowy.objects.order_by('-barcode').all()#) #.all()# oder_by('barcode')
    if request.method == 'POST':
        myform = fileform(request.POST, request.FILES)
        if myform.is_valid():
            myplik = plik()
            myplik.path = myform.cleaned_data["path"]
            myplik.save()


            readxls(myplik)
        else:
            print("NOK")
            myform=fileform()

    context['indeksy']= indeksy
    
    return render(request,url,context)

@csrf_exempt 
def przeslij_zlicz(request, *args, **kwargs):
    barcode=""
    nazwa=""
    ilosc=0
    print(request.POST)
    
    if request.method == 'POST':
        try:
            barcode = request.POST['barcode']
            skaner = request.POST['skaner']
            nazwa = request.POST['nazwa']

            ilosc = request.POST['ilosc']
            try:
                ilosc_int = int(ilosc)
            except:
                return HttpResponse("zla ilosc!")
            indeks.objects.create(barcode=barcode,nazwa=nazwa,skaner=skaner,ilosc=ilosc_int)
            return HttpResponse("dodano {}".format(barcode))
        except ValueError as a:
            return HttpResponse("blad {}".format(a))
            
    return HttpResponse("Brak dzialan")
@csrf_exempt 
def przeslij_inw(request, *args, **kwargs):
    from datetime import datetime
    from django.utils import timezone
    import pytz
    barcode=""
    nazwa=""
    ilosc=0
    if request.method == 'POST':
        try:
            barcode = request.POST['barcode']
            skaner = request.POST['skaner']
            ind = indeks_bazowy.objects.get(barcode=barcode)
            print(ind.ilosc)

            ilosc = request.POST['ilosc']
            try:
                ilosc_int = int(ilosc)
            except:
                return HttpResponse("zla ilosc!")
            ind.ilosc = ind.ilosc+ilosc_int
            ind.data_skan = datetime.now(pytz.timezone('America/New_York'))
            ind.ilosc_last=ilosc_int
            print(ind.ilosc)
            ind.save()
            return HttpResponse("dodano {}".format(barcode))
        except ValueError as a:
            return HttpResponse("blad {}".format(a))
            
    return HttpResponse("Brak dzialan")
def zerowanie_indeks(request):
    indeks.objects.all().delete()
    return HttpResponseRedirect("/mtinwentura/mtinw_pc")
def zerowanie_indeks_bazowy(request):
    indeks_bazowy.objects.all().delete()
    return HttpResponseRedirect("/mtinwentura/mtinw_pc2")


def export_xlsx(request, *args, **kwargs):
    #url='mtinwentura/export_xlsx.html'
    #context={}
    from openpyxl import Workbook,load_workbook
    from openpyxl.utils import get_column_letter
    #import os

    #path_dst=os.path.join("c:","inw","inwentura.xlsx")
    path_dst="inwentura.xlsx"
    wb_dst = Workbook()
    ws_dst = wb_dst.active
    ws_dst.title="Inwentura"

    dst_row = 1
    for i in indeks.objects.all():

         ws_dst.cell(dst_row,column=1).value=i.barcode
         ws_dst.cell(dst_row,column=2).value=i.nazwa
         ws_dst.cell(dst_row,column=3).value=str(i.data)
         dst_row += 1
    wb_dst.save(path_dst)
    ##
    filename = 'inwentura.xlsx'
    response = HttpResponse(open(filename, 'rb').read())
    response['Content-Type'] = 'mimetype/submimetype'
    response['Content-Disposition'] = 'attachment; filename=inwentura.xlsx'
    ##
    #return render(request,url,context)
    return response

def export_xlsx_bazowe(request, *args, **kwargs):
    #url='mtinwentura/export_xlsx_bazowe.html'
    #context={}
    from openpyxl import Workbook,load_workbook
    from openpyxl.utils import get_column_letter
    #import os

    #path_dst=os.path.join("c:","inw","inwentura_bazowe.xlsx")
    path_dst="inwentura_bazowa.xlsx"
    wb_dst = Workbook()
    ws_dst = wb_dst.active
    ws_dst.title="Inwentura_bazowa"

    dst_row = 1
    ws_dst.cell(dst_row,column=1).value="MATERIAŁ"
    ws_dst.cell(dst_row,column=2).value="KRÓTKI TEKST MATERIAŁU"
    ws_dst.cell(dst_row,column=3).value="KRÓTKI TEKST MATERIAŁU"
    ws_dst.cell(dst_row,column=4).value="SPIS Z NATURY"
    ws_dst.cell(dst_row,column=5).value="MIEJSCE SKŁADOWANIA"
    ws_dst.cell(dst_row,column=6).value="OSTATNIE SKANOWANIE"


    dst_row += 1
    for i in indeks_bazowy.objects.all():

         ws_dst.cell(dst_row,column=1).value=i.barcode
         ws_dst.cell(dst_row,column=2).value=i.nazwa
         ws_dst.cell(dst_row,column=3).value=i.nieogr_wyk
         ws_dst.cell(dst_row,column=4).value=i.ilosc
         ws_dst.cell(dst_row,column=5).value=i.lokalizacja
         if i.data_skan!=None:
            ws_dst.cell(dst_row,column=6).value=str(i.data_skan)[:-13]


         #ws_dst.cell(dst_row,column=3).value=str(i.data)
         dst_row += 1
    wb_dst.save(path_dst)
    filename = 'inwentura_bazowa.xlsx'
    response = HttpResponse(open(filename, 'rb').read())
    response['Content-Type'] = 'mimetype/submimetype'
    response['Content-Disposition'] = 'attachment; filename=inwentura_bazowa.xlsx'
    #return render(request,url,context)
    return response

def kopiuj_bazowe(request, *args, **kwargs):
    indeks_bazowy.objects.all().delete
    for ind in indeks.objects.order_by('barcode').distinct('barcode'):
        indeks_bazowy.objects.create(nazwa=ind.nazwa,barcode=ind.barcode)



    return HttpResponseRedirect("/mtinwentura/mtinw_pc")


def rap_xlsx_bazowe(request):
    filename = 'inwentura_bazowa.xlsx'
    response = HttpResponse(open(filename, 'rb').read())
    response['Content-Type'] = 'mimetype/submimetype'
    response['Content-Disposition'] = 'attachment; filename=inwentura_bazowa.xlsx'
    #response = HttpResponse(output,content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #response['Content-Disposition'] = 'attachment; filename=%s' % filename
    #return response
    return  HttpResponseRedirect("/mtinwentura/mtinw_pc")



def rap_xlsx(request):
    filename = 'inwentura.xlsx'
    response = HttpResponse(open(filename, 'rb').read())
    response['Content-Type'] = 'mimetype/submimetype'
    response['Content-Disposition'] = 'attachment; filename=inwentura.xlsx'
    #response = HttpResponse(output,content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #response['Content-Disposition'] = 'attachment; filename=%s' % filename
    #return response
    return  HttpResponseRedirect("/mtinwentura/mtinw_pc")







#    context = {}
#    url='mtinwentura/export_xlsx.html'
#    return render(request,url,context)

