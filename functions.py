
from .models import plik,indeks_bazowy
import os
import re
import codecs
import smtplib
def readxls(plik):
    from openpyxl import Workbook,load_workbook
    from openpyxl.utils import get_column_letter

    path=plik.path
    wb = load_workbook(path)
    ws = wb.active
    max_row = ws.max_row
    ###WYWALAMY!!!!
    indeks_bazowy.objects.all().delete()
    for row in ws.iter_rows(1,max_row):
        #indeks_bazowy.objects.create()
        try:

            #print(row[0].value) ##Barcode
            #print(row[1].value) # Nazwa
            #print(row[4].value) # Nazwa

            indeks_bazowy.objects.create(barcode=row[0].value,nazwa=row[1].value,lokalizacja=row[4].value,nieogr_wyk=row[2].value)

        except:
            print("Niepoprawny wiersz")


    return None

